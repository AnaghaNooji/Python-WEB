#!/usr/bin/python

import openpyxl
import json
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from ansible.module_utils.basic import AnsibleModule


ANSIBLE_METADATA = {
    'metadata_version': '1.1',
    'status': ['preview'],
    'supported_by': 'Ansible_Automation_Team'
}

DOCUMENTATION = r'''
module: json_to_excel
short_description: 
   Convert json file(.json) to excel file format(.xlsx) 
version_added: "2.9.16"
description:
- Convert an appointed json file to a excel file. The keys for each host recorded in the json file must be lower case. The value specified in 'categories' and 'color_column' is case insensitive.
options:
  categories:
    description:
    - "Lists the column categories shown on the first row"
    required: true
  categories_change:
    description:
    - "The actual column name we want in the sheet"
    required: false
  path:
    description:
    - "The path to the json file. Either 'path' or 'content' needs to be specified."
    required: false
  json_content:
    description:
    - "The content in dictionary format passed to be recorded in excel file. Either 'path' or 'content' needs to be specified."
    required: false
  color_rule:
    description:
    - This option defines the rule for the specified columns to have color. It only supports green, red and yellow colors so far. It is a list and its each element must be either a list or the value null. The parameters for a rule of filling color are as following.
      1. column - required<true>. example<"column':' ['column_name1', 'column_name2', ...]>. It specifies the columns names which will follow the rule of filling color defined here.
      2. value - required<true>. example<[['value_for_green1', 'value_for_green2', ...], ['value_for_red1', 'value_for_red2', ...], ['value_for_yellow1', 'value_for_yellow2', ...]]>. It contains 3 colors, green, red and yellow in order. If no value to fill in color, then specify it as null. Each element of the value option is either a list or null. The length must be 3.
      3. operator - required<false>. It is either 'In' or 'NotIn'. If this option is not defined, it is 'In' by default. 'In' means if the call value is in the value list of option 'value', then corresponding color will be filled in this cell, and vice versa for 'NotIn'.
      4. range - required<false>. example<"column_name">. Only one column name as a string is required for this 'range' value. If this option is defined, then only the rows which has applied the color for the column 'column_name' will be taken into account for the other columns to have color based on the color rule. 
    required: false
  excel_file:
    description:
    - "The path to the target excel file. If not specified, the file will be named as 'result.xlxs' in the current directory"
    required: false
  sheet_name:
    description:
    - "The sheet name. It is 'Result' by default. It will replace the sheet if the same sheet name exists."
    required: false
'''

EXAMPLES = r'''
    - name: Excel conversion dedup status to .xlsx file
      json_to_excel:
        categories: ["storage_name","vserver","volume_name","dedup_status","changelog_percent"]
        json_content: "{{ sis_json_content }}"
        color_rule:
        - column: ["dedup_status"] 
          value: [["Success"], ["Failure"], null]
        excel_file: "{{ playbook_dir }}/dedup_status.xlsx"
        sheet_name: "Result"
'''

RETURN = r'''
changed:
    description: if the transfer of json file to excel file succeeds or not.
    type: bool
    returned: always
failed:
    description: if the transfer fails or not.
    type: bool
    returned: always
total_num:
    description: total host number.
    type: int
    returned: always
fail_num:
    description: The host number which has been marked in red color for the first column name in the 'color_rule' setting.
    type: int
    returned: always
output:
    description: a nesting dict containing 'changed', 'failed', 'total_num' and 'fail_num'.
    type: dict
    returned: always
    Example:
    "output": {
        "changed": true,
        "fail_num": 2,
        "failed": false,
        "total_num": 2
    }
'''

EXTENSION = '.xlsx'

    
def main():
    module_args = dict(
        categories=dict(required=True, type='list'),
        categories_change=dict(required=False, default={}, type='dict'),
        path=dict(required=False, type='str'),
        json_content=dict(required=False, type='dict'),
        color_rule=dict(required=False, default=[], type='list'),
        excel_file=dict(required=False, default='output.xlsx', type='str'),
        sheet_name=dict(required=False, default='Output', type='str')
    )

    module = AnsibleModule(
        argument_spec=module_args,
        supports_check_mode=False
    )

    result = json_to_excel(module)

    module.exit_json(**result)
    
    
def json_to_excel(module):
    categories = module.params['categories']
    categories_change = module.params['categories_change']
    json_file = module.params['path']
    color_rule = module.params['color_rule']
    excel_file = module.params['excel_file']
    json_content = module.params['json_content']
    sheet_name = module.params['sheet_name']

    result = dict(
        changed=False,
    )

    if json_content and json_file:
        module.fail_json(msg="Both 'json_file' and 'json_content' are specified. Only one can be the input.")
        
        
    color_num = 3 # green, red, yellow
    if color_rule:
        for attr in color_rule:
            if set(attr.keys()) - {'column', 'operator', 'value', 'range'}:
                module.fail_json(msg="An unknown option for the module 'json_excel_file'. Please use the module correctly.")

            operator = attr.get("operator")
            if isinstance(operator, str) and operator not in ['In', 'NotIn']:
                module.fail_json(msg="Only 'In' and 'NotIn' are allowed for the option 'operator'. Please correct it and try it again.")

            value = attr.get("value")
            if (not isinstance(value, list)) or (isinstance(value, list) and len(value) != color_num):
                module.fail_json(msg=("The 'value' of 'color_rule' must be a list with %d elements." % color_num))

            crange = attr.get("range")
            if crange is not None and not isinstance(crange, str):
                module.fail_json(msg="The 'range' only accepts one column name as a string.")

    
    data = {}
    
    if json_content:
        data = json_content
    elif json_file:
        with open(json_file) as obj:
            for line in obj.readlines():
                dict1 = {}
                try:
                     dict1 = json.loads(line)
                except ValueError:
                     module.fail_json(msg="The json line does not contain a valid json string.")
                
                data.update(dict1) 
            
            
    try:
        wb = load_workbook(excel_file)
    except FileNotFoundError:
        wb = Workbook()
        
    ws1 = wb.active
    sheet_yes = False
    
    for sheet in wb:
        if sheet.title == sheet_name:
            wb.remove(sheet)
            continue

        if sheet.max_row == 1 and sheet.max_column == 1 and not sheet.cell(1, 1).value:
            if sheet_yes:
                wb.remove(sheet)
            else:
                sheet.delete_rows(1)
                ws1 = sheet
                sheet_yes = True
                
    if sheet_yes:
        ws1.title = sheet_name
    else:
        ws1 = wb.create_sheet(sheet_name)
        
    if categories[0].lower() not in ["host", "ip"]:
        categories.insert(0, "Host")
        
    categories_low = list(map(lambda x: x.lower() if isinstance(x, str) else x, categories[1:]))
 
    if categories_change:
        for ch in categories_change.keys():
            if ch in categories:
                categories[categories.index(ch)] = categories_change[ch]

    ws1.append(categories) 
    
    success_num, unsup_num = 0, 0

    for host, val_dict in data.items():
        # only allow the type of every val of dict is the same
        hostname = "-"
        if "hostname" in categories_low and "hostname" in val_dict:
            hostname = val_dict["hostname"]
            val_dict.pop("hostname")

        input_wait = row_input(host, hostname, categories_low, val_dict, ws1, True)
        if input_wait:
            ws1.append([host] + ['-'] * len(categories_low))

        val_unsup = val_dict.get("unsupported")
        if val_unsup == "F":
            unsup_num += 1
        elif val_unsup == "-":
            success_num += 1
            
    font = Font(bold=True, color="ffffff")
    align = Alignment(horizontal='left', vertical='center')
    color_fg = ["6CC24A", "ff0000", "F9F048"]
    fill_color = [PatternFill("solid", fgColor=color_fg[i]) for i in range(0, color_num)]
    fill_blue = PatternFill("solid", fgColor="5F249F") 

    for col in range(1, ws1.max_column + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 15
        
    for col in ws1.columns:
    # col is a tuple
        for row in range(0, ws1.max_row):
            if row == 0:
                col[row].font = font
                col[row].fill = fill_blue
            col[row].alignment = align

    # fill color for the value cells
    for attr in color_rule:
        color_col = attr.get("column")
        if color_col:
            color_col = list(map(lambda x: x.lower() if isinstance(x, str) else x, color_col))
            color_rng = attr.get("range")

            color_val = attr.get("value")
            for ival in range(0, color_num):
                color_val[ival] = list(map(lambda x: x.lower() if isinstance(x, str) else x, color_val[ival])) if color_val[ival] is not None else None

            color_opr = attr.get("operator")
            if not color_opr:
                color_opr = "In"

            color_col_index = []
            color_col_oneindex = 0
            color_rows_oneindex = set()
            for i in color_col:
                if i in categories_low:
                    index_n = categories_low.index(i) + 2
                    if color_rng is not None and color_rng.lower() == i:
                        color_col_oneindex = index_n
                        # make sure the range_static column index is always in position index 0
                        color_col_index.insert(0, index_n)
                    else:
                        color_col_index += [index_n]

            rows_range = range(2, ws1.max_row + 1)
            for icol in color_col_index:
                for irow in rows_range:
                    cell = ws1.cell(irow, icol)
                    val = cell.value

                    for icolor in range(0, color_num):
                        val_condition = False

                        if color_opr == "In":
                            val_condition = bool(color_val[icolor] and val.lower() in color_val[icolor])
                        elif  color_opr == "NotIn":
                            val_condition = bool(color_val[icolor] and val.lower() not in color_val[icolor])

                        if val_condition:
                            cell.fill = fill_color[icolor]
                            if color_col_oneindex == icol:
                                color_rows_oneindex.add(irow)

                    if color_col_oneindex == icol and irow == ws1.max_row:
                        rows_range = color_rows_oneindex

    ws1.alignment = align

    wb.save(excel_file)

    result['changed'] = True
    result['fail_num'] = unsup_num
    result['total_num'] = success_num + unsup_num

    return result

# attributes should be located at the same level
# row_wait -- status when waiting for the 1st row input, once it's input,it's turned to False.
# categories still contains "hostname" if "hostname" is from the intput
def row_input(host, hostname, categories, val, ws, row_wait=True):
    row_data = [host] + ['-'] * len(categories)

    if isinstance(val, dict):
        # find out the attributes level
        if set(val.keys()).intersection(set(categories)):
            for key, k_val in val.items():
                if key in categories:
                    row_data[categories.index(key) + 1] = str(k_val) if k_val is not None else ''

            row_data[0] = '' if not row_wait else host
            if "hostname" in categories:
                row_data[categories.index("hostname") + 1] = '' if not row_wait else hostname
            ws.append(row_data)
            row_wait = False
        else:
            for key, k_val in val.items():
                row_wait = row_input(host, hostname, categories, k_val, ws, row_wait)

    return row_wait
        
###main file   
    
if __name__ == '__main__':
    
    main()




